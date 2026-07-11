from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUTPUTS = ROOT / "outputs"
EXCEL_PATH = OUTPUTS / "complaints_dashboard.xlsx"

FILES = {
    "Monthly Trend": OUTPUTS / "monthly_complaints.csv",
    "Top Categories": OUTPUTS / "top_categories.csv",
    "Top Businesses": OUTPUTS / "top_businesses.csv",
    "Quarterly Trend": OUTPUTS / "quarterly_category_trend.csv",
    "Risk Summary": OUTPUTS / "risk_summary.csv",
}

def read_csv(path):
    if not path.exists():
        raise FileNotFoundError(f"Missing file: {path}")
    return pd.read_csv(path)

def write_workbook():
    monthly = read_csv(FILES["Monthly Trend"])
    categories = read_csv(FILES["Top Categories"])
    businesses = read_csv(FILES["Top Businesses"])
    quarterly = read_csv(FILES["Quarterly Trend"])
    risk = read_csv(FILES["Risk Summary"])

    total_complaints = int(monthly["total_complaints"].sum()) if "total_complaints" in monthly.columns else 0
    top_category = categories.iloc[0]["category_clean"] if len(categories) else "Not available"
    top_category_count = int(categories.iloc[0]["total_complaints"]) if len(categories) else 0
    top_business = businesses.iloc[0]["business_clean"] if len(businesses) else "Not available"
    top_business_count = int(businesses.iloc[0]["total_complaints"]) if len(businesses) else 0
    highest_risk = risk.iloc[0]["category_clean"] if len(risk) else "Not available"
    highest_risk_score = risk.iloc[0]["risk_score"] if len(risk) else 0

    summary = pd.DataFrame({
        "metric": [
            "Total complaints",
            "Top complaint category",
            "Top category complaint count",
            "Top business",
            "Top business complaint count",
            "Highest risk category",
            "Highest risk score"
        ],
        "value": [
            total_complaints,
            top_category,
            top_category_count,
            top_business,
            top_business_count,
            highest_risk,
            highest_risk_score
        ]
    })

    with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Dashboard", index=False, startrow=2)
        monthly.to_excel(writer, sheet_name="Monthly Trend", index=False)
        categories.to_excel(writer, sheet_name="Top Categories", index=False)
        businesses.to_excel(writer, sheet_name="Top Businesses", index=False)
        quarterly.to_excel(writer, sheet_name="Quarterly Trend", index=False)
        risk.to_excel(writer, sheet_name="Risk Summary", index=False)

def format_sheet(ws):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_fill = PatternFill("solid", fgColor="1F4E78")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="D9D9D9")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = header_fill

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    if ws.title == "Dashboard":
        ws["A1"] = "AI Evidence Auditor Dashboard"
        ws["A1"].font = white_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A1:B1")
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 45

def add_charts(wb):
    ws_monthly = wb["Monthly Trend"]
    ws_categories = wb["Top Categories"]
    ws_businesses = wb["Top Businesses"]
    ws_risk = wb["Risk Summary"]
    ws_dash = wb["Dashboard"]

    if ws_monthly.max_row > 1:
        chart = LineChart()
        chart.title = "Monthly Complaints Trend"
        chart.y_axis.title = "Total Complaints"
        chart.x_axis.title = "Month"
        data = Reference(ws_monthly, min_col=2, min_row=1, max_row=ws_monthly.max_row)
        cats = Reference(ws_monthly, min_col=1, min_row=2, max_row=ws_monthly.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 16
        ws_dash.add_chart(chart, "D3")

    if ws_categories.max_row > 1:
        chart = BarChart()
        chart.title = "Top Complaint Categories"
        chart.y_axis.title = "Total Complaints"
        chart.x_axis.title = "Category"
        data = Reference(ws_categories, min_col=2, min_row=1, max_row=min(ws_categories.max_row, 11))
        cats = Reference(ws_categories, min_col=1, min_row=2, max_row=min(ws_categories.max_row, 11))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 16
        ws_dash.add_chart(chart, "D20")

    if ws_businesses.max_row > 1:
        chart = BarChart()
        chart.title = "Top Businesses by Complaint Count"
        chart.y_axis.title = "Total Complaints"
        chart.x_axis.title = "Business"
        data = Reference(ws_businesses, min_col=2, min_row=1, max_row=min(ws_businesses.max_row, 11))
        cats = Reference(ws_businesses, min_col=1, min_row=2, max_row=min(ws_businesses.max_row, 11))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 16
        ws_dash.add_chart(chart, "D37")

    if ws_risk.max_row > 1 and "risk_score" in [cell.value for cell in ws_risk[1]]:
        chart = BarChart()
        chart.title = "Highest Risk Categories"
        chart.y_axis.title = "Risk Score"
        chart.x_axis.title = "Category"
        data = Reference(ws_risk, min_col=5, min_row=1, max_row=min(ws_risk.max_row, 11))
        cats = Reference(ws_risk, min_col=1, min_row=2, max_row=min(ws_risk.max_row, 11))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 16
        ws_dash.add_chart(chart, "D54")

def polish_workbook():
    wb = load_workbook(EXCEL_PATH)

    for ws in wb.worksheets:
        format_sheet(ws)
        ws.freeze_panes = "A2"

    add_charts(wb)

    ws = wb["Dashboard"]
    ws["A12"] = "Analyst Notes"
    ws["A12"].font = Font(bold=True)
    ws["A13"] = "This dashboard summarises complaint volume, category trends, repeated businesses, and risk-ranked complaint areas using SQL outputs from the cleaned dataset."
    ws["A14"] = "The next stage adds GenAI-generated executive summaries that must be validated against SQL evidence before use."
    ws.merge_cells("A13:B13")
    ws.merge_cells("A14:B14")
    ws["A13"].alignment = Alignment(wrap_text=True)
    ws["A14"].alignment = Alignment(wrap_text=True)

    wb.save(EXCEL_PATH)

def main():
    write_workbook()
    polish_workbook()
    print("Excel dashboard created")
    print(EXCEL_PATH)

if __name__ == "__main__":
    main()